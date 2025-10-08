from playwright.sync_api import sync_playwright
import time
import psycopg2
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime


def login():
    """Login to Duetto Analytics application"""

    # Ask user for Duetto IP with port
    duetto_url = input('\nEnter Duetto IP address with port (e.g., http://192.168.10.232:3000/): ').strip()

    # Validate URL is not empty
    if not duetto_url:
        print('[ERROR] URL cannot be empty')
        return

    # Add http:// if not present
    if not duetto_url.startswith('http://') and not duetto_url.startswith('https://'):
        duetto_url = 'http://' + duetto_url

    # Ensure trailing slash
    if not duetto_url.endswith('/'):
        duetto_url += '/'

    print('\n' + '='*70)
    print('TESTING: Login')
    print('='*70)
    print(f'[INFO] Duetto URL: {duetto_url}')

    # Start Playwright
    with sync_playwright() as p:
        # Launch browser
        browser = p.chromium.launch(headless=False, slow_mo=500)
        context = browser.new_context(viewport={'width': 1920, 'height': 1080})
        page = context.new_page()

        # Navigate to login page
        print(f'[INFO] Navigating to: {duetto_url}')

        try:
            page.goto(duetto_url)
            page.wait_for_load_state('networkidle')
            time.sleep(2)
        except Exception as e:
            print(f'[ERROR] Failed to navigate to login page: {str(e)}')
            browser.close()
            return

        # Fill in credentials
        print('[INFO] Entering credentials (admin/admin)...')
        page.fill('input[name="user_name"]', 'admin')
        page.fill('input[name="password"]', 'admin')

        # Click login button
        print('[INFO] Clicking login button...')
        page.click('button[type="submit"]')
        page.wait_for_load_state('networkidle')
        time.sleep(2)

        # Check if login was successful
        current_url = page.url

        # Give it more time to load
        time.sleep(3)
        current_url = page.url

        # Try to find dashboard indicator (button or URL change)
        is_logged_in = False

        # Check 1: URL doesn't contain login
        if 'login' not in current_url.lower():
            is_logged_in = True

        # Check 2: Try to find Dashboard button (means we're logged in)
        try:
            dashboard_btn = page.get_by_role("button", name="Dashboard")
            if dashboard_btn.count() > 0:
                is_logged_in = True
        except:
            pass

        if is_logged_in:
            print('\n[SUCCESS] Login Successful')
            print(f'[INFO] Current URL: {current_url}')
            print('='*70 + '\n')
        else:
            print('\n[FAIL] Login Failed')
            print(f'[INFO] Current URL: {current_url}')
            print('='*70 + '\n')

        # Keep browser open
        input('\nPress Enter to close browser...')
        browser.close()


def data_fetch():
    """Fetch data from PostgreSQL database and store in Excel"""

    print('\n' + '='*70)
    print('FUNCTION: Data Fetch from Database')
    print('='*70)

    # Database connection details (from .env file)
    DB_HOST = '192.168.10.232'
    DB_PORT = 5432
    DB_NAME = 'duetto_analytics_db'
    DB_USER = 'postgres'
    DB_PASSWORD = 'postgrespw'

    try:
        # Connect to PostgreSQL
        print(f'[INFO] Connecting to database at {DB_HOST}:{DB_PORT}...')
        conn = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            database=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD
        )
        print('[SUCCESS] Connected to database')

        cursor = conn.cursor()

        # Fetch all data from Devices table
        print('\n[STEP 1] Fetching data from Devices table...')
        cursor.execute('SELECT * FROM public."Devices"')

        # Get column names
        devices_columns = [desc[0] for desc in cursor.description]
        devices_data = cursor.fetchall()

        print(f'[OK] Retrieved {len(devices_data)} records from Devices table')

        # Create DataFrame for Devices
        devices_df = pd.DataFrame(devices_data, columns=devices_columns)

        # Convert timezone-aware datetime columns to timezone-naive
        for col in devices_df.columns:
            if pd.api.types.is_datetime64_any_dtype(devices_df[col]):
                if devices_df[col].dt.tz is not None:
                    devices_df[col] = devices_df[col].dt.tz_localize(None)

        # Fetch all data from DeviceSensorMasters table
        print('\n[STEP 2] Fetching data from DeviceSensorMasters table...')
        cursor.execute('SELECT * FROM public."DeviceSensorMasters"')

        # Get column names
        sensors_columns = [desc[0] for desc in cursor.description]
        sensors_data = cursor.fetchall()

        print(f'[OK] Retrieved {len(sensors_data)} records from DeviceSensorMasters table')

        # Create DataFrame for DeviceSensorMasters
        sensors_df = pd.DataFrame(sensors_data, columns=sensors_columns)

        # Convert timezone-aware datetime columns to timezone-naive
        for col in sensors_df.columns:
            if pd.api.types.is_datetime64_any_dtype(sensors_df[col]):
                if sensors_df[col].dt.tz is not None:
                    sensors_df[col] = sensors_df[col].dt.tz_localize(None)

        # Close database connection
        cursor.close()
        conn.close()
        print('\n[INFO] Database connection closed')

        # Save to Excel
        print('\n[STEP 3] Saving data to Excel...')
        excel_path = os.path.join(os.path.dirname(__file__), 'excel_container.xlsx')

        # Write both sheets
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
            devices_df.to_excel(writer, sheet_name='list of devices', index=False)
            sensors_df.to_excel(writer, sheet_name='DeviceSensorMasters', index=False)

        print(f'[SUCCESS] Data saved to: {excel_path}')
        print(f'  - Sheet 1: "list of devices" ({len(devices_df)} rows)')
        print(f'  - Sheet 2: "DeviceSensorMasters" ({len(sensors_df)} rows)')

        print('='*70 + '\n')

    except psycopg2.Error as e:
        print(f'[ERROR] Database error: {e}')
        print('='*70 + '\n')
    except Exception as e:
        print(f'[ERROR] Unexpected error: {e}')
        print('='*70 + '\n')


def parse_time_to_seconds(time_str):
    """Convert time string like '2min' to seconds"""
    time_str = time_str.strip().lower()

    if 'min' in time_str:
        return int(time_str.replace('min', '').strip()) * 60
    elif 'sec' in time_str:
        return int(time_str.replace('sec', '').strip())
    else:
        # Default to minutes if no unit specified
        return int(time_str) * 60


def get_test_configuration():
    """Get test configuration from user"""
    print('\n' + '='*70)
    print('TEST CONFIGURATION')
    print('='*70)

    duetto_url = input('\nEnter Duetto IP address with port (e.g., http://192.168.10.232:3000/): ').strip()

    # Validate URL
    if not duetto_url:
        print('[ERROR] URL cannot be empty')
        return None

    # Add http:// if not present
    if not duetto_url.startswith('http://') and not duetto_url.startswith('https://'):
        duetto_url = 'http://' + duetto_url

    # Ensure trailing slash
    if not duetto_url.endswith('/'):
        duetto_url += '/'

    iterations = int(input('\nHow many times to run tests? (e.g., 10): ').strip())

    interval_input = input('\nInterval between test runs? (e.g., 1min, 2min, 10min): ').strip()
    interval_seconds = parse_time_to_seconds(interval_input)

    refresh_input = input('\nBrowser page refresh interval? (e.g., 5min, 10min): ').strip()
    refresh_seconds = parse_time_to_seconds(refresh_input)

    print('\n[CONFIG] Test Configuration:')
    print(f'  - URL: {duetto_url}')
    print(f'  - Iterations: {iterations}')
    print(f'  - Test Interval: {interval_seconds}s ({interval_input})')
    print(f'  - Refresh Interval: {refresh_seconds}s ({refresh_input})')

    return {
        'duetto_url': duetto_url,
        'iterations': iterations,
        'interval_seconds': interval_seconds,
        'refresh_seconds': refresh_seconds,
        'interval_display': interval_input,
        'refresh_display': refresh_input
    }


def wait_with_countdown(seconds):
    """Wait with countdown display"""
    if seconds <= 0:
        return

    print(f'\n[WAIT] Waiting {seconds} seconds until next run...')
    for remaining in range(seconds, 0, -1):
        mins, secs = divmod(remaining, 60)
        print(f'\rTime remaining: {mins:02d}:{secs:02d}', end='', flush=True)
        time.sleep(1)
    print('\n')


def should_refresh_browser(last_refresh_time, refresh_interval):
    """Check if browser should be refreshed"""
    return (time.time() - last_refresh_time) >= refresh_interval


def refresh_browser(page):
    """Refresh browser page and navigate to dashboard"""
    print('\n[REFRESH] Refreshing browser page...')
    try:
        # Click Dashboard button to ensure we're on the right page
        dashboard_btn = page.get_by_role("button", name="Dashboard")
        if dashboard_btn.count() > 0:
            dashboard_btn.click()
            page.wait_for_load_state('networkidle')
            time.sleep(2)
        else:
            # Fallback to page reload
            page.reload()
            page.wait_for_load_state('networkidle')
            time.sleep(2)
        print('[SUCCESS] Page refreshed')
    except Exception as e:
        print(f'[ERROR] Failed to refresh page: {str(e)}')
        raise


def perform_login(page, duetto_url):
    """Perform login on the given page"""
    print(f'[INFO] Navigating to: {duetto_url}')

    page.goto(duetto_url)
    page.wait_for_load_state('networkidle')
    time.sleep(2)

    print('[INFO] Entering credentials (admin/admin)...')
    page.fill('input[name="user_name"]', 'admin')
    page.fill('input[name="password"]', 'admin')

    print('[INFO] Clicking login button...')
    page.click('button[type="submit"]')
    page.wait_for_load_state('networkidle')
    time.sleep(3)

    # Check if login was successful
    current_url = page.url
    is_logged_in = False

    if 'login' not in current_url.lower():
        is_logged_in = True

    try:
        dashboard_btn = page.get_by_role("button", name="Dashboard")
        if dashboard_btn.count() > 0:
            is_logged_in = True
    except:
        pass

    if is_logged_in:
        print('[SUCCESS] Login Successful')
        print(f'[INFO] Current URL: {current_url}')
        return True
    else:
        print('[FAIL] Login Failed')
        print(f'[INFO] Current URL: {current_url}')
        return False


def print_final_summary(results, exception_count, start_time):
    """Print final test summary"""
    end_time = time.time()
    duration_seconds = int(end_time - start_time)
    duration_mins = duration_seconds // 60
    duration_secs = duration_seconds % 60

    print('\n' + '#'*70)
    print('# FINAL TEST SUMMARY')
    print('#'*70)

    total_runs = len(results)
    passed = sum(1 for r in results if 'PASS' in r['status'])
    failed = total_runs - passed

    print(f'\nTotal test runs: {total_runs}')
    print(f'Passed: {passed}')
    print(f'Failed: {failed}')
    print(f'Exceptions handled: {exception_count}')
    print(f'Total duration: {duration_mins}m {duration_secs}s')

    print('\nDetailed Results:')
    print('-' * 70)
    for result in results:
        status_symbol = '✓' if 'PASS' in result['status'] else '✗'
        print(f"{status_symbol} Run {result['run']}: {result['status']}")
        print(f"   - Device Count: {'PASS' if result['device_count'] else 'FAIL'}")
        print(f"   - Sensor Sequence: {'PASS' if result['sensor_sequence'] else 'FAIL'}")

    print('#'*70 + '\n')


def build_fail_description(device_count_result, sensor_sequence_result):
    """Build detailed fail description from verification results"""
    fail_lines = []

    # Check device count failures
    if not device_count_result['pass']:
        fail_lines.append('DEVICE COUNT VERIFICATION FAILED:')
        fail_lines.append(f"  Expected: {device_count_result['expected_count']} devices")
        fail_lines.append(f"  Actual: {device_count_result['actual_count']} devices")

        if device_count_result['missing_devices']:
            fail_lines.append(f"  Missing devices: {', '.join(device_count_result['missing_devices'])}")

        if device_count_result['extra_devices']:
            fail_lines.append(f"  Extra devices: {', '.join(device_count_result['extra_devices'])}")

        fail_lines.append('')  # Empty line separator

    # Check sensor sequence failures
    if not sensor_sequence_result['pass']:
        fail_lines.append('SENSOR SEQUENCE VERIFICATION FAILED:')
        fail_lines.append(f"  Devices tested: {sensor_sequence_result['devices_tested']}")
        fail_lines.append(f"  Devices passed: {sensor_sequence_result['devices_passed']}")
        fail_lines.append(f"  Devices failed: {sensor_sequence_result['devices_failed']}")
        fail_lines.append('')

        # Add details for each failed device
        if sensor_sequence_result['failure_details']:
            fail_lines.append('  Failed Devices:')
            for idx, failure in enumerate(sensor_sequence_result['failure_details'], 1):
                fail_lines.append(f"  {idx}. {failure['device_name']}:")
                for mismatch in failure['sensor_mismatches']:
                    fail_lines.append(f"     - {mismatch}")

    return '\n'.join(fail_lines) if fail_lines else ''


def capture_failure_screenshot(page, test_run):
    """Capture screenshot on test failure"""
    # Create screenshots directory if it doesn't exist
    screenshots_dir = os.path.join(os.path.dirname(__file__), 'screenshots')
    os.makedirs(screenshots_dir, exist_ok=True)

    # Generate screenshot filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    screenshot_filename = f'test_run_{test_run}_fail_{timestamp}.png'
    screenshot_path = os.path.join(screenshots_dir, screenshot_filename)

    # Capture full page screenshot
    page.screenshot(path=screenshot_path, full_page=True)

    print(f'[SCREENSHOT] Saved: {screenshot_path}')

    # Return relative path for Excel
    return f'screenshots/{screenshot_filename}'


def initialize_excel_report():
    """Initialize or load Excel report file"""
    report_path = os.path.join(os.path.dirname(__file__), 'Report.xlsx')

    # Check if file exists
    if os.path.exists(report_path):
        # Load existing workbook
        wb = load_workbook(report_path)
        if 'Test Results' in wb.sheetnames:
            ws = wb['Test Results']
        else:
            ws = wb.create_sheet('Test Results')
    else:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Results'

    # Clear existing data and set headers
    ws.delete_rows(1, ws.max_row)

    # Set headers
    headers = ['Test Run', 'Pass/Fail', 'Fail Description', 'References']
    ws.append(headers)

    # Format headers
    for col in range(1, 5):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='top')

    # Set column widths
    ws.column_dimensions['A'].width = 12  # Test Run
    ws.column_dimensions['B'].width = 12  # Pass/Fail
    ws.column_dimensions['C'].width = 80  # Fail Description
    ws.column_dimensions['D'].width = 40  # References

    # Save
    wb.save(report_path)
    print(f'[EXCEL] Initialized report: {report_path}')

    return report_path


def write_test_result_to_excel(test_run, pass_fail, fail_description, screenshot_path):
    """Write test result to Excel report"""
    report_path = os.path.join(os.path.dirname(__file__), 'Report.xlsx')

    # Load workbook
    wb = load_workbook(report_path)
    ws = wb['Test Results']

    # Append new row
    row_data = [test_run, pass_fail, fail_description, screenshot_path]
    ws.append(row_data)

    # Format the new row
    row_num = ws.max_row

    # Center align Test Run and Pass/Fail columns
    ws.cell(row_num, 1).alignment = Alignment(horizontal='center', vertical='top')
    ws.cell(row_num, 2).alignment = Alignment(horizontal='center', vertical='top')

    # Wrap text for Fail Description
    ws.cell(row_num, 3).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row_num].height = None  # Auto-height

    # Format References
    ws.cell(row_num, 4).alignment = Alignment(vertical='top')

    # Set Pass/Fail cell color
    if pass_fail == 'PASS':
        ws.cell(row_num, 2).font = Font(color='008000', bold=True)  # Green
    else:
        ws.cell(row_num, 2).font = Font(color='FF0000', bold=True)  # Red

    # Save
    wb.save(report_path)
    print(f'[EXCEL] Updated report: Test Run {test_run} - {pass_fail}')


def verify_device_count(page):
    """Verify Air Quality Stations tab has correct number of devices"""

    print('\n' + '='*70)
    print('TEST CASE 1: Device Quantity Verification')
    print('='*70)

    # STEP 1: Load expected data from Excel
    print('\n[STEP 1] Loading expected data from Excel...')
    excel_path = os.path.join(os.path.dirname(__file__), 'excel_container.xlsx')

    devices_df = pd.read_excel(excel_path, sheet_name='list of devices')

    # Filter AQS devices (device_type_id = 3, 4, 5)
    aqs_devices = devices_df[devices_df['device_type_id'].isin([3, 4, 5])].copy()
    aqs_devices = aqs_devices.sort_values('id')

    print(f'[INFO] Expected AQS devices: {len(aqs_devices)}')
    for _, device in aqs_devices.iterrows():
        print(f'  - Device ID {device["id"]}: {device["device_name"]} (type_id: {device["device_type_id"]})')

    # STEP 2: Navigate to Dashboard
    print('\n[STEP 2] Navigating to Dashboard...')

    # Check if already on dashboard, if not navigate
    try:
        dashboard_btn = page.get_by_role("button", name="Dashboard")
        if dashboard_btn.count() > 0:
            print('[INFO] Already logged in, clicking Dashboard...')
            dashboard_btn.click()
            page.wait_for_load_state('networkidle')
            time.sleep(2)
    except:
        print('[ERROR] Not logged in or Dashboard button not found')
        return False

    # STEP 3: Verify we're on Air Quality Stations tab (Tab 0 - should be default)
    print('\n[STEP 3] Verifying Air Quality Stations tab...')

    # Wait for table to load
    try:
        page.wait_for_selector('table tbody tr', timeout=10000)
        time.sleep(2)  # Extra wait for data to populate
    except:
        print('[ERROR] Table not found or not loaded')
        return False

    # STEP 4: Count devices in frontend
    print('\n[STEP 4] Counting devices in frontend...')

    # Get all table rows
    table_rows = page.locator('table tbody tr')
    actual_device_count = table_rows.count()

    print(f'[INFO] Found devices in frontend: {actual_device_count}')

    # STEP 5: Extract and verify device names
    print('\n[STEP 5] Verifying device list...')

    expected_device_names = set(aqs_devices['device_name'].tolist())
    found_device_names = set()

    for row_idx in range(actual_device_count):
        row = table_rows.nth(row_idx)
        try:
            device_name_cell = row.locator('td').nth(1)
            device_name_text = device_name_cell.inner_text().split('\n')[0].strip()
            found_device_names.add(device_name_text)
            print(f'  ✓ Found: {device_name_text}')
        except Exception as e:
            print(f'  ✗ Error reading device at row {row_idx}: {str(e)}')

    # STEP 6: Compare results
    print('\n[STEP 6] Comparison Results...')

    missing_devices = expected_device_names - found_device_names
    extra_devices = found_device_names - expected_device_names

    count_match = actual_device_count == len(aqs_devices)
    list_match = len(missing_devices) == 0 and len(extra_devices) == 0

    if missing_devices:
        print(f'[WARN] Missing devices: {missing_devices}')

    if extra_devices:
        print(f'[WARN] Extra devices not in database: {extra_devices}')

    # STEP 7: Summary
    print('\n' + '='*70)
    print('TEST CASE 1 SUMMARY')
    print('='*70)
    print(f'Expected device count: {len(aqs_devices)}')
    print(f'Actual device count: {actual_device_count}')
    print(f'Device count match: {"✓ PASS" if count_match else "✗ FAIL"}')
    print(f'Device list match: {"✓ PASS" if list_match else "✗ FAIL"}')

    overall_pass = count_match and list_match

    if overall_pass:
        print('\n[OVERALL RESULT] ✓ PASS - Device quantity verification passed!')
    else:
        print('\n[OVERALL RESULT] ✗ FAIL - Device quantity verification failed')

    print('='*70 + '\n')

    # Return detailed result dictionary
    result = {
        'pass': overall_pass,
        'expected_count': len(aqs_devices),
        'actual_count': actual_device_count,
        'missing_devices': list(missing_devices),
        'extra_devices': list(extra_devices)
    }
    return result


def verify_sensor_sequence(page):
    """Verify sensor sequence for each device on Air Quality Stations tab"""

    print('\n' + '='*70)
    print('TEST CASE 2: Sensor Sequence Verification')
    print('='*70)

    # STEP 1: Load expected data from Excel
    print('\n[STEP 1] Loading expected data from Excel...')
    excel_path = os.path.join(os.path.dirname(__file__), 'excel_container.xlsx')

    devices_df = pd.read_excel(excel_path, sheet_name='list of devices')
    sensors_df = pd.read_excel(excel_path, sheet_name='DeviceSensorMasters')

    # Filter AQS devices (device_type_id = 3, 4, 5)
    aqs_devices = devices_df[devices_df['device_type_id'].isin([3, 4, 5])].copy()
    aqs_devices = aqs_devices.sort_values('id')

    # Prepare expected sensor sequences for each device
    expected_data = {}
    for _, device in aqs_devices.iterrows():
        device_id = device['id']
        device_name = device['device_name']

        # Get sensors for this device
        device_sensors = sensors_df[sensors_df['device_id'] == device_id].copy()

        # Filter configured sensors (is_configured != 0)
        # If all are 0 or NaN, show all sensors
        if device_sensors['is_configured'].notna().any():
            configured_sensors = device_sensors[device_sensors['is_configured'] != 0]
            if len(configured_sensors) > 0:
                device_sensors = configured_sensors

        # Sort by priority (priority=0 goes to end)
        device_sensors = device_sensors.sort_values(
            by='priority',
            key=lambda x: x.replace(0, 999)  # Replace 0 with high number to push to end
        )

        # Take first 10 sensors
        device_sensors = device_sensors.head(10)

        # Build expected sensor list
        expected_sensors = []
        for idx, (_, sensor) in enumerate(device_sensors.iterrows(), 1):
            sensor_info = {
                'position': idx,
                'location': sensor['location'],
                'config_type': int(sensor['config_type']) if pd.notna(sensor['config_type']) else None,
                'sensor_type': int(sensor['sensor_type']) if pd.notna(sensor['sensor_type']) else None,
                'priority': int(sensor['priority']) if pd.notna(sensor['priority']) else 0
            }
            expected_sensors.append(sensor_info)

        expected_data[device_id] = {
            'device_name': device_name,
            'expected_sensor_count': len(expected_sensors),
            'sensors': expected_sensors
        }

    print(f'[INFO] Testing sensor sequences for {len(expected_data)} devices')

    # STEP 2: Navigate to Dashboard
    print('\n[STEP 2] Navigating to Dashboard...')

    # Check if already on dashboard, if not navigate
    try:
        dashboard_btn = page.get_by_role("button", name="Dashboard")
        if dashboard_btn.count() > 0:
            print('[INFO] Already logged in, clicking Dashboard...')
            dashboard_btn.click()
            page.wait_for_load_state('networkidle')
            time.sleep(2)
    except:
        print('[ERROR] Not logged in or Dashboard button not found')
        return False

    # STEP 3: Verify we're on Air Quality Stations tab (Tab 0 - should be default)
    print('\n[STEP 3] Verifying Air Quality Stations tab...')

    # Wait for table to load
    try:
        page.wait_for_selector('table tbody tr', timeout=10000)
        time.sleep(2)  # Extra wait for data to populate
    except:
        print('[ERROR] Table not found or not loaded')
        return False

    # STEP 4: Get table rows
    table_rows = page.locator('table tbody tr')
    actual_device_count = table_rows.count()

    # STEP 5: Verify sensor sequence for each device
    print('\n[STEP 4] Verifying sensor sequences...')
    print('-' * 70)

    passed_devices = 0
    failed_devices = 0
    failure_details = []  # Store detailed failure info

    for row_idx in range(actual_device_count):
        row = table_rows.nth(row_idx)

        # Extract device name from Location column (2nd column, index 1)
        try:
            device_name_cell = row.locator('td').nth(1)
            device_name_text = device_name_cell.inner_text().split('\n')[0].strip()

            print(f'\n[Device {row_idx + 1}] {device_name_text}')

            # Find this device in expected data
            device_id = None
            for dev_id, dev_data in expected_data.items():
                if dev_data['device_name'] == device_name_text:
                    device_id = dev_id
                    break

            if device_id is None:
                print(f'[WARN] Device "{device_name_text}" not found in expected data')
                failed_devices += 1
                continue

            expected_sensors = expected_data[device_id]['sensors']
            print(f'Expected sensors: {len(expected_sensors)}')

            # Extract sensors from columns 3-12 (Sensor 1-10)
            sensor_match_count = 0
            sensor_mismatch_count = 0
            sensor_mismatches = []  # Store mismatches for this device

            for sensor_idx, expected_sensor in enumerate(expected_sensors):
                position = expected_sensor['position']
                expected_location = expected_sensor['location']
                expected_config_type = expected_sensor['config_type']

                # Column index: 0=Product, 1=Location, 2=Sensor1, 3=Sensor2, etc.
                sensor_column_idx = 1 + position  # +1 because location is at index 1

                try:
                    sensor_cell = row.locator('td').nth(sensor_column_idx)
                    sensor_text = sensor_cell.inner_text().strip()

                    # Check if sensor location matches
                    if expected_location and expected_location in sensor_text:
                        print(f'  ✓ Sensor {position}: {expected_location} (config_type: {expected_config_type})')
                        sensor_match_count += 1
                    else:
                        mismatch_msg = f'Sensor {position}: Expected "{expected_location}", but found "{sensor_text[:50]}"'
                        print(f'  ✗ {mismatch_msg}')
                        sensor_mismatch_count += 1
                        sensor_mismatches.append(mismatch_msg)

                except Exception as e:
                    error_msg = f'Sensor {position}: Could not read sensor data - {str(e)}'
                    print(f'  ✗ {error_msg}')
                    sensor_mismatch_count += 1
                    sensor_mismatches.append(error_msg)

            # Check remaining columns should be empty
            for empty_position in range(len(expected_sensors) + 1, 11):
                sensor_column_idx = 1 + empty_position
                try:
                    sensor_cell = row.locator('td').nth(sensor_column_idx)
                    sensor_text = sensor_cell.inner_text().strip()
                    if sensor_text and 'Not Installed' not in sensor_text:
                        print(f'  [WARN] Sensor {empty_position}: Expected empty, but found content')
                except:
                    pass  # Empty is expected

            # Determine pass/fail for this device
            if sensor_mismatch_count == 0:
                print(f'[PASS] Sensor sequence verified ({sensor_match_count}/{len(expected_sensors)})')
                passed_devices += 1
            else:
                print(f'[FAIL] Sensor sequence mismatch ({sensor_match_count} matches, {sensor_mismatch_count} mismatches)')
                failed_devices += 1
                # Store failure details
                failure_details.append({
                    'device_name': device_name_text,
                    'sensor_mismatches': sensor_mismatches
                })

        except Exception as e:
            error_msg = f'Failed to process device row {row_idx}: {str(e)}'
            print(f'[ERROR] {error_msg}')
            failed_devices += 1
            failure_details.append({
                'device_name': f'Device at row {row_idx + 1}',
                'sensor_mismatches': [error_msg]
            })

    # STEP 6: Summary
    print('\n' + '='*70)
    print('TEST CASE 2 SUMMARY')
    print('='*70)
    print(f'Total devices tested: {passed_devices + failed_devices}')
    print(f'Passed: {passed_devices}')
    print(f'Failed: {failed_devices}')

    overall_pass = failed_devices == 0

    if overall_pass:
        print('\n[OVERALL RESULT] ✓ PASS - All sensor sequences verified!')
    else:
        print('\n[OVERALL RESULT] ✗ FAIL - Some sensor sequences failed')

    print('='*70 + '\n')

    # Return detailed result dictionary
    result = {
        'pass': overall_pass,
        'devices_tested': passed_devices + failed_devices,
        'devices_passed': passed_devices,
        'devices_failed': failed_devices,
        'failure_details': failure_details
    }
    return result


def test_case_one():
    """Enhanced Test Case 1: Verify device quantity and sensor sequence with configurable iterations"""

    print('\n' + '#'*70)
    print('# TEST CASE 1: Air Quality Stations Verification')
    print('# Enhanced with Loop, Interval, and Exception Handling')
    print('#'*70 + '\n')

    # Step 1: Fetch latest data from database
    print('[WORKFLOW] Step 1: Fetching latest data from database...')
    data_fetch()

    # Step 2: Get test configuration
    print('\n[WORKFLOW] Step 2: Getting test configuration...')
    config = get_test_configuration()

    if config is None:
        print('[ERROR] Invalid configuration')
        return

    # Initialize tracking
    results = []
    exception_count = 0
    start_time = time.time()
    last_refresh_time = start_time

    # Initialize Excel report
    print('\n[WORKFLOW] Initializing Excel report...')
    initialize_excel_report()

    # Step 3: Launch browser and login
    print('\n[WORKFLOW] Step 3: Launching browser and logging in...')

    with sync_playwright() as p:
        # Launch browser
        browser = p.chromium.launch(headless=False, slow_mo=500)
        context = browser.new_context(viewport={'width': 1920, 'height': 1080})
        page = context.new_page()

        # Initial login
        login_success = perform_login(page, config['duetto_url'])

        if not login_success:
            browser.close()
            print('[ERROR] Initial login failed. Exiting.')
            return

        # Step 4: Run test loop
        print('\n[WORKFLOW] Step 4: Starting test iterations...')
        print(f"[INFO] Will run {config['iterations']} iterations with {config['interval_display']} interval\n")

        for iteration in range(1, config['iterations'] + 1):
            print('\n' + '='*70)
            print(f'TEST RUN {iteration}/{config["iterations"]}')
            print('='*70)

            try:
                # Check if browser needs refresh
                if should_refresh_browser(last_refresh_time, config['refresh_seconds']):
                    print(f'[INFO] {config["refresh_display"]} passed since last refresh')
                    refresh_browser(page)
                    last_refresh_time = time.time()

                # Run device count verification
                print(f'\n[RUN {iteration}] Running device quantity verification...')
                device_count_result = verify_device_count(page)

                # Run sensor sequence verification
                print(f'\n[RUN {iteration}] Running sensor sequence verification...')
                sensor_sequence_result = verify_sensor_sequence(page)

                # Determine overall pass/fail
                overall_pass = device_count_result['pass'] and sensor_sequence_result['pass']

                # Build fail description and capture screenshot if failed
                fail_description = ''
                screenshot_path = ''
                if not overall_pass:
                    fail_description = build_fail_description(device_count_result, sensor_sequence_result)
                    screenshot_path = capture_failure_screenshot(page, iteration)

                # Write to Excel report
                write_test_result_to_excel(
                    test_run=iteration,
                    pass_fail='PASS' if overall_pass else 'FAIL',
                    fail_description=fail_description,
                    screenshot_path=screenshot_path
                )

                # Record results for final summary
                results.append({
                    'run': iteration,
                    'device_count': device_count_result['pass'],
                    'sensor_sequence': sensor_sequence_result['pass'],
                    'status': 'PASS' if overall_pass else 'FAIL'
                })

                print(f'\n[RUN {iteration}] Result: {"✓ PASS" if overall_pass else "✗ FAIL"}')

            except Exception as e:
                print(f'\n[EXCEPTION] Error in run {iteration}: {str(e)}')
                print('[RECOVERY] Attempting to restart browser and re-login...')

                exception_count += 1

                # Close current browser
                try:
                    browser.close()
                except:
                    pass

                # Small wait before restart
                time.sleep(2)

                # Restart browser and re-login
                try:
                    browser = p.chromium.launch(headless=False, slow_mo=500)
                    context = browser.new_context(viewport={'width': 1920, 'height': 1080})
                    page = context.new_page()

                    login_success = perform_login(page, config['duetto_url'])

                    if not login_success:
                        print('[ERROR] Re-login failed after exception')

                        # Write failure to Excel
                        write_test_result_to_excel(
                            test_run=iteration,
                            pass_fail='FAIL',
                            fail_description='Re-login failed after exception',
                            screenshot_path=''
                        )

                        results.append({
                            'run': iteration,
                            'device_count': False,
                            'sensor_sequence': False,
                            'status': 'FAILED (login error)'
                        })
                        continue

                    last_refresh_time = time.time()

                    # Retry the test
                    print(f'\n[RETRY] Re-running test {iteration} after recovery...')

                    device_count_result = verify_device_count(page)
                    sensor_sequence_result = verify_sensor_sequence(page)

                    overall_pass = device_count_result['pass'] and sensor_sequence_result['pass']

                    # Build fail description and capture screenshot if failed
                    fail_description = ''
                    screenshot_path = ''
                    if not overall_pass:
                        fail_description = build_fail_description(device_count_result, sensor_sequence_result)
                        screenshot_path = capture_failure_screenshot(page, iteration)

                    # Write to Excel report
                    write_test_result_to_excel(
                        test_run=iteration,
                        pass_fail='PASS (after retry)' if overall_pass else 'FAIL (after retry)',
                        fail_description=fail_description,
                        screenshot_path=screenshot_path
                    )

                    # Record results for final summary
                    results.append({
                        'run': iteration,
                        'device_count': device_count_result['pass'],
                        'sensor_sequence': sensor_sequence_result['pass'],
                        'status': 'PASS (after retry)' if overall_pass else 'FAIL (after retry)'
                    })

                    print(f'\n[RUN {iteration}] Result after retry: {"✓ PASS" if overall_pass else "✗ FAIL"}')

                except Exception as retry_e:
                    print(f'[ERROR] Retry also failed: {str(retry_e)}')

                    # Write failure to Excel
                    write_test_result_to_excel(
                        test_run=iteration,
                        pass_fail='FAIL',
                        fail_description=f'Exception during retry: {str(retry_e)}',
                        screenshot_path=''
                    )

                    results.append({
                        'run': iteration,
                        'device_count': False,
                        'sensor_sequence': False,
                        'status': 'FAILED (exception)'
                    })

            # Wait interval before next run (except for last iteration)
            if iteration < config['iterations']:
                wait_with_countdown(config['interval_seconds'])

        # Step 5: Print final summary
        print_final_summary(results, exception_count, start_time)

        # Keep browser open
        input('\nPress Enter to close browser and exit...')
        browser.close()

    print('\n[WORKFLOW] Test Case 1 completed!\n')


if __name__ == '__main__':
    # Run test case one
    test_case_one()
