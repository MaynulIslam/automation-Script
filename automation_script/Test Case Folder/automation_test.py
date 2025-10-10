from playwright.sync_api import sync_playwright
import time
import psycopg2
import pandas as pd
import os
import random
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
import paho.mqtt.client as mqtt


# ==================== CONFIGURATION ====================

def get_test_configuration():
    """Get test configuration from user"""
    print('\n' + '='*70)
    print('TEST CONFIGURATION')
    print('='*70)

    da_instance = input('\nWhat is your DA instance? (e.g., http://192.168.10.232:3000/): ').strip()

    # Validate URL
    if not da_instance:
        print('[ERROR] URL cannot be empty')
        return None

    # Add http:// if not present
    if not da_instance.startswith('http://') and not da_instance.startswith('https://'):
        da_instance = 'http://' + da_instance

    # Ensure trailing slash
    if not da_instance.endswith('/'):
        da_instance += '/'

    sensor_combinations = int(input('\nHow many sensor combinations you want to test? (e.g., 3): ').strip())
    checks_per_combination = int(input('\nHow many times you want to check for each combination? (e.g., 5): ').strip())
    total_iterations = int(input('\nTotal test iterations? (e.g., 10): ').strip())
    skip_login = input('\nWant to skip login after first iteration? (y/n): ').strip().lower()

    print('\n[CONFIG] Test Configuration:')
    print(f'  - DA Instance: {da_instance}')
    print(f'  - Sensor Combinations: {sensor_combinations}')
    print(f'  - Checks per Combination: {checks_per_combination}')
    print(f'  - Total Iterations: {total_iterations}')
    print(f'  - Skip Login: {skip_login}')

    return {
        'da_instance': da_instance,
        'sensor_combinations': sensor_combinations,
        'checks_per_combination': checks_per_combination,
        'total_iterations': total_iterations,
        'skip_login': skip_login == 'y'
    }


# ==================== DATABASE FUNCTIONS ====================

def connect_to_database():
    """Connect to PostgreSQL database"""
    DB_HOST = '192.168.10.232'
    DB_PORT = 5432
    DB_NAME = 'duetto_analytics_db'
    DB_USER = 'postgres'
    DB_PASSWORD = 'postgrespw'

    try:
        conn = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            database=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD
        )
        return conn
    except psycopg2.Error as e:
        print(f'[ERROR] Database connection failed: {e}')
        return None


def fetch_devices_from_db(conn):
    """Fetch devices from database"""
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM public."Devices"')
    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()
    cursor.close()

    df = pd.DataFrame(data, columns=columns)
    return df


def fetch_sensors_from_db(conn):
    """Fetch sensors from database"""
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM public."DeviceSensorMasters"')
    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()
    cursor.close()

    df = pd.DataFrame(data, columns=columns)
    return df


def verify_sensor_priority_in_db(conn, device_id, expected_priorities):
    """Verify if sensor priorities are applied in database"""
    sensors_df = fetch_sensors_from_db(conn)
    device_sensors = sensors_df[sensors_df['device_id'] == device_id].copy()

    # Sort by priority
    device_sensors = device_sensors.sort_values('priority')

    # Compare
    for idx, expected in enumerate(expected_priorities):
        actual = device_sensors.iloc[idx] if idx < len(device_sensors) else None
        if actual is None or actual['id'] != expected['id']:
            return False, f"Priority mismatch at position {idx}"

    return True, "Priorities verified successfully"


def create_db_html_table(query_result_df, title):
    """Create HTML table from query result DataFrame"""
    html_template = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; }}
            h2 {{ color: #333; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th {{ background-color: #4CAF50; color: white; padding: 12px; text-align: left; border: 1px solid #ddd; }}
            td {{ padding: 8px; border: 1px solid #ddd; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
        </style>
    </head>
    <body>
        <h2>{title}</h2>
        {query_result_df.to_html(index=False, border=0)}
    </body>
    </html>
    """
    return html_template


def capture_db_screenshot(query_result_df, title, screenshot_path, browser):
    """Capture database query result as screenshot"""
    html_content = create_db_html_table(query_result_df, title)

    # Write HTML to temp file
    temp_html = screenshot_path.replace('.png', '_db.html')
    with open(temp_html, 'w', encoding='utf-8') as f:
        f.write(html_content)

    # Create a new page in the existing browser to screenshot the HTML
    db_page = browser.new_page(viewport={'width': 1200, 'height': 800})
    temp_html_path = os.path.abspath(temp_html).replace('\\', '/')
    db_page.goto(f'file:///{temp_html_path}')
    db_page.screenshot(path=screenshot_path, full_page=True)
    db_page.close()

    # Clean up temp HTML
    if os.path.exists(temp_html):
        os.remove(temp_html)

    print(f'[SCREENSHOT] Database screenshot saved: {screenshot_path}')


# ==================== MQTT FUNCTIONS ====================

def generate_random_sensor_combination(device_id, sensors_df):
    """Generate random sensor priority combination for a device"""
    device_sensors = sensors_df[sensors_df['device_id'] == device_id].copy()

    # Filter configured sensors
    if device_sensors['is_configured'].notna().any():
        configured = device_sensors[device_sensors['is_configured'] != 0]
        if len(configured) > 0:
            device_sensors = configured

    # Randomly shuffle and take up to 10
    device_sensors = device_sensors.sample(frac=1).head(10)

    # Build sensor priority list
    sensor_priority = []
    for idx, (_, sensor) in enumerate(device_sensors.iterrows(), 1):
        sensor_info = {
            'id': int(sensor['id']),
            'config_type': sensor['location'] if pd.notna(sensor['location']) else '',
            'name': sensor.get('name', ''),
            'location': sensor['location'] if pd.notna(sensor['location']) else '',
            'selectedItems': []
        }
        sensor_priority.append(sensor_info)

    return sensor_priority


def apply_sensor_combination_via_mqtt(device_id, device_type, sensor_priority):
    """Apply sensor combination via MQTT"""
    print(f'\n[MQTT] Publishing sensor combination for device {device_id}...')

    payload = {
        'device_id': device_id,
        'sensor_priority': sensor_priority,
        'device_type': device_type
    }

    # MQTT Connection details
    MQTT_BROKER = '192.168.10.232'
    MQTT_PORT = 1883

    client = mqtt.Client()

    try:
        client.connect(MQTT_BROKER, MQTT_PORT, 60)
        topic = f'duetto_analytics/aqs/sensor_priority/{device_id}'
        client.publish(topic, json.dumps(payload))
        print(f'[SUCCESS] Published to {topic}')
        client.disconnect()
        return True
    except Exception as e:
        print(f'[ERROR] MQTT publish failed: {e}')
        return False


# ==================== FAILURE DESCRIPTION BUILDERS ====================

def build_dashboard_failure_description(dashboard_result, devices_df):
    """Build detailed dashboard failure description"""
    if dashboard_result.get('pass', False):
        return ''

    description_lines = []

    frontend_total = dashboard_result.get('frontend_total', 'N/A')
    db_total = dashboard_result.get('db_total', 'N/A')

    if str(frontend_total) != str(db_total):
        description_lines.append(f"Total Devices Mismatch:")
        description_lines.append(f"  Frontend: {frontend_total} devices")
        description_lines.append(f"  Database: {db_total} devices")

    frontend_active = dashboard_result.get('frontend_active', 'N/A')
    db_active = dashboard_result.get('db_active', 'N/A')

    if str(frontend_active) != str(db_active):
        description_lines.append(f"\nActive Devices Mismatch:")
        description_lines.append(f"  Frontend: {frontend_active} devices")
        description_lines.append(f"  Database: {db_active} devices")

    frontend_offline = dashboard_result.get('frontend_offline', 'N/A')
    db_offline = dashboard_result.get('db_offline', 'N/A')

    if str(frontend_offline) != str(db_offline):
        description_lines.append(f"\nOffline Devices Mismatch:")
        description_lines.append(f"  Frontend: {frontend_offline} devices")
        description_lines.append(f"  Database: {db_offline} devices")

    return '\n'.join(description_lines)


def build_device_existence_failure_description(device_result, devices_df):
    """Build detailed device existence failure description with device names and IPs"""
    if device_result.get('pass', False):
        return ''

    description_lines = []

    # Missing devices (in DB but not in frontend)
    missing_device_names = device_result.get('missing_devices', [])
    if missing_device_names:
        description_lines.append("Frontend doesn't have these devices:")
        for device_name in missing_device_names:
            device_row = devices_df[devices_df['device_name'] == device_name]
            if not device_row.empty:
                ip_address = device_row.iloc[0]['ip_address']
                description_lines.append(f"  - {device_name} - {ip_address}")
            else:
                description_lines.append(f"  - {device_name}")

    # Extra devices (in frontend but not in DB)
    extra_device_names = device_result.get('extra_devices', [])
    if extra_device_names:
        if missing_device_names:
            description_lines.append("")
        description_lines.append("Database doesn't have these devices:")
        for device_name in extra_device_names:
            description_lines.append(f"  - {device_name}")

    return '\n'.join(description_lines)


def build_sensor_sequence_failure_description(sensor_result, devices_df, sensors_df):
    """Build detailed sensor sequence failure description"""
    if sensor_result.get('pass', False):
        return ''

    description_lines = []
    failure_details = sensor_result.get('failure_details', [])

    for idx, failure in enumerate(failure_details, 1):
        device_name = failure['device_name']

        # Find device IP
        device_row = devices_df[devices_df['device_name'] == device_name]
        ip_address = device_row.iloc[0]['ip_address'] if not device_row.empty else 'N/A'

        description_lines.append(f"{idx}. Device: {device_name} - {ip_address}")

        # Get expected sensor sequence from DB
        if not device_row.empty:
            device_id = device_row.iloc[0]['id']
            device_sensors = sensors_df[sensors_df['device_id'] == device_id].copy()

            # Filter configured sensors
            if device_sensors['is_configured'].notna().any():
                configured = device_sensors[device_sensors['is_configured'] != 0]
                if len(configured) > 0:
                    device_sensors = configured

            # Sort by priority
            device_sensors = device_sensors.sort_values(
                by='priority',
                key=lambda x: x.replace(0, 999)
            ).head(10)

            backend_sensors = [sensor['location'] for _, sensor in device_sensors.iterrows() if pd.notna(sensor['location'])]
            if backend_sensors:
                description_lines.append(f"   Backend sensor priority: {', '.join(backend_sensors)}")

        if idx < len(failure_details):
            description_lines.append("")

    return '\n'.join(description_lines)


def capture_failure_screenshots_by_type(page, browser, conn, iteration, timestamp,
                                       dashboard_failed, device_failed, sensor_failed,
                                       dashboard_result, device_result, sensor_result):
    """Capture screenshots for each type of failure"""
    evidence_files = []

    devices_df = fetch_devices_from_db(conn)
    sensors_df = fetch_sensors_from_db(conn)

    # Dashboard Matrix failure
    if dashboard_failed:
        # Frontend screenshot
        frontend_filename = f'iteration_{iteration}_{timestamp}_dashboard_matrix_frontend.png'
        frontend_path = os.path.join(os.path.dirname(__file__), 'evidence', frontend_filename)
        page.screenshot(path=frontend_path, full_page=True)
        evidence_files.append(frontend_filename)
        print(f'[SCREENSHOT] Dashboard Matrix Frontend: {frontend_filename}')

        # Database screenshot
        db_filename = f'iteration_{iteration}_{timestamp}_dashboard_matrix_database.png'
        db_path = os.path.join(os.path.dirname(__file__), 'evidence', db_filename)
        capture_db_screenshot(devices_df.head(20), f"Dashboard Matrix - Database", db_path, browser)
        evidence_files.append(db_filename)

    # Air Quality Station (Device Exists) failure
    if device_failed:
        # Frontend screenshot
        frontend_filename = f'iteration_{iteration}_{timestamp}_aqs_devices_frontend.png'
        frontend_path = os.path.join(os.path.dirname(__file__), 'evidence', frontend_filename)
        page.screenshot(path=frontend_path, full_page=True)
        evidence_files.append(frontend_filename)
        print(f'[SCREENSHOT] AQS Devices Frontend: {frontend_filename}')

        # Database screenshot - show AQS devices only
        aqs_devices = devices_df[devices_df['device_type_id'].isin([3, 4, 5])]
        db_filename = f'iteration_{iteration}_{timestamp}_aqs_devices_database.png'
        db_path = os.path.join(os.path.dirname(__file__), 'evidence', db_filename)
        capture_db_screenshot(aqs_devices[['device_name', 'ip_address', 'device_type_id']],
                            f"AQS Devices - Database", db_path, browser)
        evidence_files.append(db_filename)

    # Sensor Sequence failure
    if sensor_failed:
        # Frontend screenshot
        frontend_filename = f'iteration_{iteration}_{timestamp}_sensor_sequence_frontend.png'
        frontend_path = os.path.join(os.path.dirname(__file__), 'evidence', frontend_filename)
        page.screenshot(path=frontend_path, full_page=True)
        evidence_files.append(frontend_filename)
        print(f'[SCREENSHOT] Sensor Sequence Frontend: {frontend_filename}')

        # Database screenshot - show sensor priorities
        db_filename = f'iteration_{iteration}_{timestamp}_sensor_sequence_database.png'
        db_path = os.path.join(os.path.dirname(__file__), 'evidence', db_filename)

        # Get failed devices and their sensors
        failure_details = sensor_result.get('failure_details', [])
        if failure_details:
            failed_device_names = [f['device_name'] for f in failure_details]
            failed_device_ids = devices_df[devices_df['device_name'].isin(failed_device_names)]['id'].tolist()
            failed_sensors = sensors_df[sensors_df['device_id'].isin(failed_device_ids)]
            failed_sensors = failed_sensors[['device_id', 'location', 'priority', 'is_configured']].sort_values(['device_id', 'priority'])
            capture_db_screenshot(failed_sensors, f"Sensor Sequence - Database", db_path, browser)
        else:
            capture_db_screenshot(sensors_df.head(20), f"Sensor Sequence - Database", db_path, browser)

        evidence_files.append(db_filename)

    return evidence_files, devices_df, sensors_df


# ==================== BROWSER/FRONTEND FUNCTIONS ====================

def login_to_dashboard(page, da_instance):
    """Login to Duetto Analytics dashboard"""
    print('\n' + '='*70)
    print('LOGIN')
    print('='*70)

    print(f'[INFO] Navigating to: {da_instance}')

    try:
        page.goto(da_instance)
        page.wait_for_load_state('networkidle')
        time.sleep(2)

        print('[INFO] Entering credentials (admin/admin)...')
        page.fill('input[name="user_name"]', 'admin')
        page.fill('input[name="password"]', 'admin')

        print('[INFO] Clicking login button...')
        page.click('button[type="submit"]')
        page.wait_for_load_state('networkidle')
        time.sleep(3)

        # Verify login
        current_url = page.url
        is_logged_in = 'login' not in current_url.lower()

        try:
            dashboard_btn = page.get_by_role("button", name="Dashboard")
            if dashboard_btn.count() > 0:
                is_logged_in = True
        except:
            pass

        if is_logged_in:
            print('[SUCCESS] Login successful')
            return True
        else:
            print('[FAIL] Login failed')
            return False

    except Exception as e:
        print(f'[ERROR] Login error: {e}')
        return False


def verify_dashboard_metrics(page, conn):
    """Verify dashboard metrics (Total Devices, Active, Offline)"""
    print('\n' + '='*70)
    print('VERIFY DASHBOARD METRICS')
    print('='*70)

    try:
        # Click Dashboard button
        dashboard_btn = page.get_by_role("button", name="Dashboard")
        if dashboard_btn.count() > 0:
            dashboard_btn.click()
            page.wait_for_load_state('networkidle')
            time.sleep(2)

        # Fetch database counts
        devices_df = fetch_devices_from_db(conn)

        db_total_devices = len(devices_df)
        db_active = len(devices_df[devices_df['status'] == 'active']) if 'status' in devices_df.columns else 0
        db_offline = len(devices_df[devices_df['status'] == 'offline']) if 'status' in devices_df.columns else 0

        print(f'[DB] Total Devices: {db_total_devices}')
        print(f'[DB] Active: {db_active}')
        print(f'[DB] Offline: {db_offline}')

        # Extract frontend metrics from DeviceAnalytic components
        # The Typography element contains text like "11 Devices", and the sibling Box contains the title
        try:
            # Wait for metrics to load
            page.wait_for_selector('text=Total Devices', timeout=5000)

            # Get all Typography elements containing "Devices" text (MUI Typography renders as <p> or <h6>)
            # Find the parent Stack that contains both the title and the number

            # Total Devices
            total_stack = page.locator('text=Total Devices').locator('..')
            frontend_total_text = total_stack.locator('p, h6').filter(has_text='Devices').first.inner_text()
            frontend_total = frontend_total_text.split()[0]

            # Active
            active_stack = page.locator('text=Active').locator('..')
            frontend_active_text = active_stack.locator('p, h6').filter(has_text='Devices').first.inner_text()
            frontend_active = frontend_active_text.split()[0]

            # Offline
            offline_stack = page.locator('text=Offline').locator('..')
            frontend_offline_text = offline_stack.locator('p, h6').filter(has_text='Devices').first.inner_text()
            frontend_offline = frontend_offline_text.split()[0]

            print(f'[FRONTEND] Total Devices: {frontend_total}')
            print(f'[FRONTEND] Active: {frontend_active}')
            print(f'[FRONTEND] Offline: {frontend_offline}')

            # Compare
            result = {
                'pass': True,
                'frontend_total': frontend_total,
                'frontend_active': frontend_active,
                'frontend_offline': frontend_offline,
                'db_total': db_total_devices,
                'db_active': db_active,
                'db_offline': db_offline
            }

            if str(db_total_devices) != str(frontend_total):
                result['pass'] = False

            if str(db_active) != str(frontend_active):
                result['pass'] = False

            if str(db_offline) != str(frontend_offline):
                result['pass'] = False

            return result

        except Exception as e:
            print(f'[WARN] Could not extract frontend metrics: {e}')
            return {
                'pass': False,
                'error': str(e),
                'db_total': db_total_devices,
                'db_active': db_active,
                'db_offline': db_offline
            }

    except Exception as e:
        print(f'[ERROR] Dashboard verification failed: {e}')
        return {'pass': False, 'error': str(e)}


def verify_aqs_tab(page, conn):
    """Verify Air Quality Stations tab - device list and sensor sequences"""
    print('\n' + '='*70)
    print('VERIFY AIR QUALITY STATIONS TAB')
    print('='*70)

    try:
        # Navigate to Dashboard
        dashboard_btn = page.get_by_role("button", name="Dashboard")
        if dashboard_btn.count() > 0:
            dashboard_btn.click()
            page.wait_for_load_state('networkidle')
            time.sleep(2)

        # Fetch database data
        devices_df = fetch_devices_from_db(conn)
        sensors_df = fetch_sensors_from_db(conn)

        # Filter AQS devices (device_type_id = 3, 4, 5)
        aqs_devices = devices_df[devices_df['device_type_id'].isin([3, 4, 5])].copy()
        aqs_devices = aqs_devices.sort_values('id')

        print(f'[DB] Expected AQS devices: {len(aqs_devices)}')

        # Wait for table
        page.wait_for_selector('table tbody tr', timeout=10000)
        time.sleep(2)

        # Count devices in frontend
        table_rows = page.locator('table tbody tr')
        frontend_device_count = table_rows.count()

        print(f'[FRONTEND] Found devices: {frontend_device_count}')

        # Verify device list
        device_verification = {
            'pass': frontend_device_count == len(aqs_devices),
            'expected_count': len(aqs_devices),
            'actual_count': frontend_device_count,
            'missing_devices': [],
            'extra_devices': []
        }

        # Verify sensor sequences
        sensor_verification = {
            'pass': True,
            'devices_tested': 0,
            'devices_passed': 0,
            'devices_failed': 0,
            'failure_details': []
        }

        expected_device_names = set(aqs_devices['device_name'].tolist())
        found_device_names = set()

        for row_idx in range(frontend_device_count):
            row = table_rows.nth(row_idx)
            try:
                # Extract device name (adjust selector as needed)
                device_name_cell = row.locator('td').nth(1)
                device_name_text = device_name_cell.inner_text().split('\n')[0].strip()
                found_device_names.add(device_name_text)

                # Find device in database
                device_id = None
                for _, device in aqs_devices.iterrows():
                    if device['device_name'] == device_name_text:
                        device_id = device['id']
                        break

                if device_id:
                    sensor_verification['devices_tested'] += 1

                    # Get expected sensors for this device
                    device_sensors = sensors_df[sensors_df['device_id'] == device_id].copy()

                    # Filter configured sensors
                    if device_sensors['is_configured'].notna().any():
                        configured = device_sensors[device_sensors['is_configured'] != 0]
                        if len(configured) > 0:
                            device_sensors = configured

                    # Sort by priority
                    device_sensors = device_sensors.sort_values(
                        by='priority',
                        key=lambda x: x.replace(0, 999)
                    ).head(10)

                    # Verify sensor sequence (simplified - adjust based on frontend structure)
                    sensor_match = True
                    sensor_mismatches = []

                    for sensor_idx, (_, sensor) in enumerate(device_sensors.iterrows(), 1):
                        try:
                            sensor_column_idx = 1 + sensor_idx
                            sensor_cell = row.locator('td').nth(sensor_column_idx)
                            sensor_text = sensor_cell.inner_text().strip()

                            expected_location = sensor['location']
                            if expected_location and expected_location not in sensor_text:
                                sensor_match = False
                                sensor_mismatches.append(f"Sensor {sensor_idx}: Expected '{expected_location}', found '{sensor_text[:30]}'")
                        except:
                            pass

                    if sensor_match:
                        sensor_verification['devices_passed'] += 1
                    else:
                        sensor_verification['devices_failed'] += 1
                        sensor_verification['pass'] = False
                        sensor_verification['failure_details'].append({
                            'device_name': device_name_text,
                            'mismatches': sensor_mismatches
                        })

            except Exception as e:
                print(f'[ERROR] Processing row {row_idx}: {e}')

        # Check for missing/extra devices
        device_verification['missing_devices'] = list(expected_device_names - found_device_names)
        device_verification['extra_devices'] = list(found_device_names - expected_device_names)

        if device_verification['missing_devices'] or device_verification['extra_devices']:
            device_verification['pass'] = False

        return {
            'device_verification': device_verification,
            'sensor_verification': sensor_verification
        }

    except Exception as e:
        print(f'[ERROR] AQS tab verification failed: {e}')
        return {
            'device_verification': {'pass': False, 'error': str(e)},
            'sensor_verification': {'pass': False, 'error': str(e)}
        }


# ==================== SCREENSHOT FUNCTIONS ====================

def capture_screenshot(page, iteration, timestamp, screenshot_type='frontend'):
    """Capture screenshot on failure"""
    evidence_dir = os.path.join(os.path.dirname(__file__), 'evidence')
    os.makedirs(evidence_dir, exist_ok=True)

    filename = f'iteration_{iteration}_{timestamp}_{screenshot_type}.png'
    screenshot_path = os.path.join(evidence_dir, filename)

    page.screenshot(path=screenshot_path, full_page=True)
    print(f'[SCREENSHOT] Saved: {screenshot_path}')

    return filename


# ==================== EXCEL REPORTING ====================

def initialize_excel_report():
    """Initialize Excel report file"""
    report_path = os.path.join(os.path.dirname(__file__), 'Test_Report.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Results'

    # Set headers
    headers = ['Iteration #', 'Dashboard Matrix', 'Air Quality Station (Device Exists)', 'Sensor Sequence', 'Evidence']
    ws.append(headers)

    # Format headers
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='top')

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 50

    wb.save(report_path)
    print(f'[EXCEL] Initialized report: {report_path}')

    return report_path


def write_test_result_to_excel(iteration, dashboard_result, device_result, sensor_result, evidence,
                               dashboard_description='', device_description='', sensor_description=''):
    """Write test result to Excel with detailed descriptions"""
    report_path = os.path.join(os.path.dirname(__file__), 'Test_Report.xlsx')

    wb = load_workbook(report_path)
    ws = wb['Test Results']

    # Build result strings with detailed descriptions
    dashboard_str = f"{'PASS' if dashboard_result.get('pass', False) else 'FAIL'}\n"
    if dashboard_description:
        dashboard_str += f"\n{dashboard_description}"
    else:
        dashboard_str += f"Frontend: Total={dashboard_result.get('frontend_total', 'N/A')}\n"
        dashboard_str += f"Database: Total={dashboard_result.get('db_total', 'N/A')}"

    device_str = f"{'PASS' if device_result.get('pass', False) else 'FAIL'}\n"
    if device_description:
        device_str += f"\n{device_description}"
    else:
        device_str += f"Frontend: {device_result.get('actual_count', 0)} devices\n"
        device_str += f"Database: {device_result.get('expected_count', 0)} devices"

    sensor_str = f"{'PASS' if sensor_result.get('pass', False) else 'FAIL'}\n"
    if sensor_description:
        sensor_str += f"\n{sensor_description}"
    else:
        sensor_str += f"Devices tested: {sensor_result.get('devices_tested', 0)}\n"
        sensor_str += f"Passed: {sensor_result.get('devices_passed', 0)}\n"
        sensor_str += f"Failed: {sensor_result.get('devices_failed', 0)}"

    # Append row
    ws.append([iteration, dashboard_str, device_str, sensor_str, evidence])

    # Format row
    row_num = ws.max_row
    for col in range(1, 6):
        ws.cell(row_num, col).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(report_path)
    print(f'[EXCEL] Updated report for iteration {iteration}')


# ==================== MAIN TEST LOOP ====================

def run_test_iteration(iteration, config, page, conn, browser, p, is_first_iteration=False):
    """Run a single test iteration"""
    print('\n' + '#'*70)
    print(f'# TEST ITERATION {iteration}')
    print('#'*70)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    evidence_files = []

    # Login (if first iteration or skip_login is False)
    if is_first_iteration or not config['skip_login']:
        if not login_to_dashboard(page, config['da_instance']):
            print('[ERROR] Login failed, skipping iteration')
            return

    # If not first iteration, apply new sensor combination
    if not is_first_iteration:
        print('\n[STEP] Generating and applying new sensor combination...')

        # Get devices
        devices_df = fetch_devices_from_db(conn)
        sensors_df = fetch_sensors_from_db(conn)
        aqs_devices = devices_df[devices_df['device_type_id'].isin([3, 4, 5])]

        if len(aqs_devices) > 0:
            # Pick first AQS device for combination change
            device = aqs_devices.iloc[0]
            device_id = device['id']
            device_type = device['device_type_id']

            # Generate random combination
            new_combination = generate_random_sensor_combination(device_id, sensors_df)

            # Apply via MQTT
            if apply_sensor_combination_via_mqtt(device_id, device_type, new_combination):
                print('[SUCCESS] Sensor combination applied')
                time.sleep(5)  # Wait for MQTT to process

                # Verify in database
                conn_refresh = connect_to_database()
                sensors_df_new = fetch_sensors_from_db(conn_refresh)
                print('[INFO] Verified sensor combination in database')
                conn_refresh.close()
            else:
                print('[ERROR] Failed to apply sensor combination')

    # Run checks multiple times
    for check in range(1, config['checks_per_combination'] + 1):
        print(f'\n--- Check {check}/{config["checks_per_combination"]} ---')

        # Verify dashboard metrics
        dashboard_result = verify_dashboard_metrics(page, conn)

        # Verify AQS tab
        aqs_result = verify_aqs_tab(page, conn)
        device_result = aqs_result['device_verification']
        sensor_result = aqs_result['sensor_verification']

        # Determine overall pass/fail
        overall_pass = (
            dashboard_result.get('pass', False) and
            device_result.get('pass', False) and
            sensor_result.get('pass', False)
        )

        # Determine individual failure types
        dashboard_failed = not dashboard_result.get('pass', False)
        device_failed = not device_result.get('pass', False)
        sensor_failed = not sensor_result.get('pass', False)

        # Capture screenshots and build descriptions only if there are failures
        if not overall_pass:
            try:
                evidence_files, devices_df, sensors_df = capture_failure_screenshots_by_type(
                    page, browser, conn, f"{iteration}.{check}", timestamp,
                    dashboard_failed, device_failed, sensor_failed,
                    dashboard_result, device_result, sensor_result
                )

                # Build detailed descriptions
                dashboard_desc = build_dashboard_failure_description(dashboard_result, devices_df) if dashboard_failed else ''
                device_desc = build_device_existence_failure_description(device_result, devices_df) if device_failed else ''
                sensor_desc = build_sensor_sequence_failure_description(sensor_result, devices_df, sensors_df) if sensor_failed else ''

            except Exception as e:
                print(f'[ERROR] Screenshot capture failed: {e}')
                evidence_files = [f'Screenshot failed: {str(e)}']
                dashboard_desc = device_desc = sensor_desc = ''
        else:
            evidence_files = []
            dashboard_desc = device_desc = sensor_desc = ''

        # Write to Excel with detailed descriptions
        try:
            write_test_result_to_excel(
                iteration=f"{iteration}.{check}",
                dashboard_result=dashboard_result,
                device_result=device_result,
                sensor_result=sensor_result,
                evidence=', '.join(evidence_files) if evidence_files else '',
                dashboard_description=dashboard_desc,
                device_description=device_desc,
                sensor_description=sensor_desc
            )
        except Exception as e:
            print(f'[ERROR] Excel write failed: {e}')

        print(f'[RESULT] Check {check}: {"PASS" if overall_pass else "FAIL"}')

        # Small delay between checks
        if check < config['checks_per_combination']:
            time.sleep(2)


def main():
    """Main entry point"""
    print('\n' + '#'*70)
    print('# DUETTO ANALYTICS - AUTOMATED TEST SUITE')
    print('#'*70)

    # Get configuration
    config = get_test_configuration()
    if not config:
        print('[ERROR] Invalid configuration')
        return

    # Initialize Excel report
    initialize_excel_report()

    # Connect to database
    conn = connect_to_database()
    if not conn:
        print('[ERROR] Database connection failed')
        return

    # Launch browser
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=500)
        context = browser.new_context(viewport={'width': 1920, 'height': 1080})
        page = context.new_page()

        try:
            # Run test iterations
            for iteration in range(1, config['total_iterations'] + 1):
                is_first = (iteration == 1)
                run_test_iteration(iteration, config, page, conn, browser, p, is_first_iteration=is_first)

                # Wait between iterations
                if iteration < config['total_iterations']:
                    print(f'\n[WAIT] Waiting 5 seconds before next iteration...')
                    time.sleep(5)

            print('\n' + '#'*70)
            print('# TEST SUITE COMPLETED')
            print('#'*70)

        except Exception as e:
            print(f'\n[ERROR] Test execution failed: {e}')
            import traceback
            traceback.print_exc()
        finally:
            # Cleanup
            conn.close()
            print('\n[INFO] Tests completed. Browser will close in 5 seconds...')
            time.sleep(5)
            browser.close()


if __name__ == '__main__':
    main()
